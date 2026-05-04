from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import bcrypt
import json
import pandas as pd
from models import db, Usuario, Producto, CarritoItem, Orden, OrdenItem
from io import BytesIO
from flask import send_file
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import ReporteETL

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui_cambiala'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///usuarios.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página'


@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Función para hashear contraseña
def hash_password(password):
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

# Función para verificar contraseña
def verify_password(password, password_hash):
    return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))

# Crear tablas y usuario admin por defecto
with app.app_context():
    db.create_all()
    # Crear usuario admin si no existe
    admin = Usuario.query.filter_by(usuario='admin').first()
    if not admin:
        admin_password = hash_password('admin123')
        admin = Usuario(
            nombre='Administrador2',
            usuario='admin',
            celular='0000000000',
            email='admin@ejemplo.com',
            password_hash=admin_password,
            es_admin=True
        )
        db.session.add(admin)
        db.session.commit()
        print("Usuario admin creado: usuario='admin', contraseña='admin123'")

# Ruta principal
@app.route('/')
def index():
    return redirect(url_for('login'))

# Endpoint de registro
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        usuario = request.form.get('usuario')
        celular = request.form.get('celular')
        email = request.form.get('email')
        password = request.form.get('password')
        
        # Verificar si el usuario o email ya existe
        usuario_existente = Usuario.query.filter_by(usuario=usuario).first()
        email_existente = Usuario.query.filter_by(email=email).first()
        
        if usuario_existente:
            flash('El nombre de usuario ya está en uso', 'error')
            return redirect(url_for('registro'))
        
        if email_existente:
            flash('El correo electrónico ya está registrado', 'error')
            return redirect(url_for('registro'))
        
        # Hashear contraseña
        password_hash = hash_password(password)
        
        # Crear nuevo usuario (por defecto no es admin)
        nuevo_usuario = Usuario(
            nombre=nombre,
            usuario=usuario,
            celular=celular,
            email=email,
            password_hash=password_hash,
            es_admin=False
        )
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        flash('Registro exitoso. Por favor inicia sesión', 'success')
        return redirect(url_for('login'))
    
    return render_template('registro.html')

# Endpoint de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        password = request.form.get('password')
        
        user = Usuario.query.filter_by(usuario=usuario).first()
        
        if user and verify_password(password, user.password_hash):
            login_user(user)
            flash(f'Bienvenido {user.nombre}!', 'success')
            
            # Redirigir según rol
            if user.es_admin:
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('tienda'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

# Perfil de usuario normal
@app.route('/perfil')
@login_required
def perfil():
    return render_template('perfil.html', usuario=current_user)


@app.route('/dashboard')
@login_required
def dashboard():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('perfil'))
    
    usuarios = Usuario.query.all()  # ← Cambiar a usuarios
    return render_template('dashboard.html', usuarios=usuarios)  # ← Cambiar a usuarios


# API Endpoint para crear usuario (desde dashboard)
@app.route('/api/usuarios', methods=['POST'])
@login_required
def crear_usuario_api():
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    data = request.json
    
    # Validar campos
    required_fields = ['nombre', 'usuario', 'celular', 'email', 'password']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'Campo {field} requerido'}), 400
    
    # Verificar si existe
    if Usuario.query.filter_by(usuario=data['usuario']).first():
        return jsonify({'error': 'Usuario ya existe'}), 400
    
    if Usuario.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email ya existe'}), 400
    
    # Hashear contraseña
    password_hash = hash_password(data['password'])
    
    nuevo_usuario = Usuario(
        nombre=data['nombre'],
        usuario=data['usuario'],
        celular=data['celular'],
        email=data['email'],
        password_hash=password_hash,
        es_admin=data.get('es_admin', False)
    )
    
    db.session.add(nuevo_usuario)
    db.session.commit()
    
    return jsonify({
        'mensaje': 'Usuario creado exitosamente',
        'usuario': {
            'id': nuevo_usuario.id,
            'nombre': nuevo_usuario.nombre,
            'usuario': nuevo_usuario.usuario,
            'email': nuevo_usuario.email
        }
    }), 201

# Endpoint para actualizar usuario
@app.route('/api/usuarios/<int:usuario_id>', methods=['PUT'])
@login_required
def actualizar_usuario_api(usuario_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    usuario = Usuario.query.get_or_404(usuario_id)
    data = request.json
    
    # Actualizar campos
    if 'nombre' in data:
        usuario.nombre = data['nombre']
    if 'usuario' in data:
        # Verificar que el nuevo usuario no exista
        if Usuario.query.filter_by(usuario=data['usuario']).first() and usuario.usuario != data['usuario']:
            return jsonify({'error': 'Nombre de usuario ya existe'}), 400
        usuario.usuario = data['usuario']
    if 'celular' in data:
        usuario.celular = data['celular']
    if 'email' in data:
        if Usuario.query.filter_by(email=data['email']).first() and usuario.email != data['email']:
            return jsonify({'error': 'Email ya existe'}), 400
        usuario.email = data['email']
    if 'password' in data and data['password']:
        usuario.password_hash = hash_password(data['password'])
    if 'es_admin' in data:
        # No permitir que se quite el admin al último administrador
        admin_count = Usuario.query.filter_by(es_admin=True).count()
        if not data['es_admin'] and admin_count == 1 and usuario.es_admin:
            return jsonify({'error': 'No puedes quitar permisos de admin al único administrador'}), 400
        usuario.es_admin = data['es_admin']
    
    db.session.commit()
    
    return jsonify({
        'mensaje': 'Usuario actualizado exitosamente',
        'usuario': {
            'id': usuario.id,
            'nombre': usuario.nombre,
            'usuario': usuario.usuario,
            'email': usuario.email,
            'es_admin': usuario.es_admin
        }
    })

# Endpoint para eliminar usuario
@app.route('/api/usuarios/<int:usuario_id>', methods=['DELETE'])
@login_required
def eliminar_usuario_api(usuario_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # No permitir eliminar al propio admin
    if usuario.id == current_user.id:
        return jsonify({'error': 'No puedes eliminar tu propio usuario'}), 400
    
    # No permitir eliminar al último administrador
    if usuario.es_admin:
        admin_count = Usuario.query.filter_by(es_admin=True).count()
        if admin_count == 1:
            return jsonify({'error': 'No puedes eliminar al único administrador'}), 400
    
    db.session.delete(usuario)
    db.session.commit()
    
    return jsonify({'mensaje': 'Usuario eliminado exitosamente'})

# Página para agregar producto
@app.route('/producto/nuevo')
@login_required
def nuevo_producto():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('tienda'))
    
    return render_template('agregar_productos.html')

# API para crear producto (recibe los datos del formulario)
@app.route('/api/productos', methods=['POST'])
@login_required
def crear_producto():
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    data = request.json
    
    # Validar campos requeridos
    if not data.get('nombre') or not data.get('precio'):
        return jsonify({'error': 'Nombre y precio son requeridos'}), 400
    
    # Crear el producto
    nuevo_producto = Producto(
        nombre=data['nombre'],
        descripcion=data.get('descripcion', ''),
        precio=float(data['precio']),
        stock=int(data.get('stock', 0)),
        categoria=data.get('categoria', ''),
        imagen_url=data.get('imagen_url', '')
    )
    
    db.session.add(nuevo_producto)
    db.session.commit()
    
    return jsonify({
        'mensaje': 'Producto creado exitosamente',
        'producto_id': nuevo_producto.id
    }), 201
    
# Gestión completa de productos (CRUD)
@app.route('/gestionar/productos')
@login_required
def gestionar_productos():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('tienda'))
    
    productos = Producto.query.all()
    return render_template('gestionar_productos.html', productos=productos)    
    
# API para editar producto
@app.route('/api/productos/<int:producto_id>', methods=['PUT'])
@login_required
def actualizar_producto(producto_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    producto = Producto.query.get_or_404(producto_id)
    data = request.json
    
    # Actualizar campos
    if 'nombre' in data:
        producto.nombre = data['nombre']
    if 'descripcion' in data:
        producto.descripcion = data['descripcion']
    if 'precio' in data:
        producto.precio = float(data['precio'])
    if 'stock' in data:
        producto.stock = int(data['stock'])
    if 'categoria' in data:
        producto.categoria = data['categoria']
    
    db.session.commit()
    
    return jsonify({'mensaje': 'Producto actualizado exitosamente'})


# API para eliminar producto
@app.route('/api/productos/<int:producto_id>', methods=['DELETE'])
@login_required
def eliminar_producto(producto_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    producto = Producto.query.get_or_404(producto_id)
    
    # Verificar que no esté en ningún carrito
    en_carrito = CarritoItem.query.filter_by(producto_id=producto_id).first()
    if en_carrito:
        return jsonify({'error': 'No se puede eliminar: producto está en un carrito'}), 400
    
    db.session.delete(producto)
    db.session.commit()
    
    return jsonify({'mensaje': 'Producto eliminado exitosamente'})

# API para obtener un producto específico
@app.route('/api/productos/<int:producto_id>', methods=['GET'])
@login_required
def obtener_producto(producto_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    producto = Producto.query.get_or_404(producto_id)
    
    return jsonify({
        'id': producto.id,
        'nombre': producto.nombre,
        'descripcion': producto.descripcion,
        'precio': producto.precio,
        'stock': producto.stock,
        'categoria': producto.categoria
    })
    
@app.route('/api/productos/buscar')
@login_required
def buscar_productos():
    query = request.args.get('q', '')
    if query:
        productos = Producto.query.filter(Producto.nombre.contains(query)).all()
    else:
        productos = Producto.query.all()
    
    return jsonify([{
        'id': p.id,
        'nombre': p.nombre,
        'descripcion': p.descripcion,
        'precio': p.precio,
        'stock': p.stock,
        'categoria': p.categoria,
        'imagen_url': p.imagen_url
    } for p in productos])    

# Obtener cantidad de items en el carrito
@app.route('/api/carrito/count')
@login_required
def carrito_count():
    count = CarritoItem.query.filter_by(usuario_id=current_user.id).count()
    return jsonify({'count': count})

# Agregar producto al carrito (API)
@app.route('/api/carrito/agregar', methods=['POST'])
@login_required
def agregar_carrito():
    try:
        data = request.json
        producto_id = data.get('producto_id')
        cantidad = data.get('cantidad', 1)
        
        # Validar que el producto existe
        producto = Producto.query.get(producto_id)
        if not producto:
            return jsonify({'error': 'Producto no encontrado'}), 404
        
        # Validar stock
        if producto.stock < cantidad:
            return jsonify({'error': f'Stock insuficiente. Solo hay {producto.stock} unidades'}), 400
        
        # Buscar si el producto ya está en el carrito del usuario
        carrito_item = CarritoItem.query.filter_by(
            usuario_id=current_user.id,
            producto_id=producto_id
        ).first()
        
        if carrito_item:
            # Si ya existe, actualizar cantidad
            nueva_cantidad = carrito_item.cantidad + cantidad
            if producto.stock < nueva_cantidad:
                return jsonify({'error': f'Stock insuficiente. Stock disponible: {producto.stock}'}), 400
            carrito_item.cantidad = nueva_cantidad
        else:
            # Si no existe, crear nuevo item
            carrito_item = CarritoItem(
                usuario_id=current_user.id,
                producto_id=producto_id,
                cantidad=cantidad
            )
            db.session.add(carrito_item)
        
        db.session.commit()
        
        return jsonify({
            'mensaje': f'✅ {cantidad}x {producto.nombre} agregado al carrito'
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
# Ver el carrito de compras
@app.route('/carrito')
@login_required
def ver_carrito():
    # Obtener todos los items del carrito del usuario actual
    items = CarritoItem.query.filter_by(usuario_id=current_user.id).all()
    
    # Calcular el total
    total = 0
    for item in items:
        total += item.producto.precio * item.cantidad
    
    return render_template('carrito.html', items=items, total=total)    

@app.route('/tienda')
@login_required
def tienda():
    # Obtener todos los productos de la base de datos
    productos = Producto.query.all()
    
    # Renderizar la plantilla pasando los productos
    return render_template('productos.html', productos=productos)

@app.route('/producto/<int:producto_id>')
@login_required
def ver_producto(producto_id):
    # Buscar el producto por su ID
    producto = Producto.query.get_or_404(producto_id)
    
    return render_template('producto_detalle.html', producto=producto)

@app.route('/api/carrito/eliminar/<int:item_id>', methods=['DELETE'])
@login_required
def eliminar_carrito_item(item_id):
    item = CarritoItem.query.get_or_404(item_id)
    
    # Verificar que el item pertenece al usuario actual
    if item.usuario_id != current_user.id:
        return jsonify({'error': 'No autorizado'}), 403
    
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'mensaje': 'Producto eliminado del carrito'})

# Logout
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada exitosamente', 'success')
    return redirect(url_for('login'))


# Procesar checkout (crear orden)
@app.route('/api/checkout', methods=['POST'])
@login_required
def procesar_checkout():
    try:
        # Obtener items del carrito
        items = CarritoItem.query.filter_by(usuario_id=current_user.id).all()
        
        if not items:
            return jsonify({'error': 'El carrito está vacío'}), 400
        
        # Calcular total
        total = 0
        for item in items:
            total += item.producto.precio * item.cantidad
        
        # Crear la orden
        orden = Orden(
            usuario_id=current_user.id,
            total=total,
            estado='pendiente'
        )
        db.session.add(orden)
        db.session.flush()  # Para obtener el ID de la orden
        
        # Crear items de la orden y actualizar stock
        for item in items:
            orden_item = OrdenItem(
                orden_id=orden.id,
                producto_id=item.producto_id,
                cantidad=item.cantidad,
                precio_unitario=item.producto.precio
            )
            db.session.add(orden_item)
            
            # Actualizar stock del producto
            item.producto.stock -= item.cantidad
        
        # Vaciar el carrito
        for item in items:
            db.session.delete(item)
        
        db.session.commit()
        
        return jsonify({
            'mensaje': '✅ Compra realizada con éxito',
            'orden_id': orden.id,
            'total': total
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
    # Ver histórico de órdenes del usuario
@app.route('/mis-ordenes')
@login_required
def mis_ordenes():
    ordenes = Orden.query.filter_by(usuario_id=current_user.id).order_by(Orden.fecha.desc()).all()
    return render_template('mis_ordenes.html', ordenes=ordenes)

# Ver detalle de una orden específica
@app.route('/orden/<int:orden_id>')
@login_required
def ver_orden(orden_id):
    orden = Orden.query.get_or_404(orden_id)
    
    # Verificar que la orden pertenece al usuario actual o es admin
    if orden.usuario_id != current_user.id and not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('tienda'))
    
    return render_template('orden_detalle.html', orden=orden)

    
# ============ ENDPOINTS ETL ============

@app.route('/dashboard/etl')
@login_required
def dashboard_etl():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('tienda'))
    
    from models import ReporteETL
    import json
    
    reporte_ventas = ReporteETL.query.filter_by(tipo_reporte='ventas').order_by(ReporteETL.fecha_reporte.desc()).first()
    reporte_productos = ReporteETL.query.filter_by(tipo_reporte='productos').order_by(ReporteETL.fecha_reporte.desc()).first()
    reporte_usuarios = ReporteETL.query.filter_by(tipo_reporte='usuarios').order_by(ReporteETL.fecha_reporte.desc()).first()
    
    ventas = json.loads(reporte_ventas.datos) if reporte_ventas else {}
    productos = json.loads(reporte_productos.datos) if reporte_productos else {}
    usuarios = json.loads(reporte_usuarios.datos) if reporte_usuarios else {}
    
    return render_template('dashboard_etl.html',
                         ventas=ventas,
                         productos=productos,
                         usuarios=usuarios,
                         ultima_actualizacion=reporte_ventas.fecha_reporte if reporte_ventas else None)

@app.route('/api/etl/ejecutar')
@login_required
def api_ejecutar_etl():
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    from etl import ejecutar_etl
    ejecutar_etl()
    
    return jsonify({'mensaje': 'ETL ejecutado correctamente'})

def aplicar_formato_excel(writer, sheet_name, columnas_ancho=None, formato_moneda=None):
    """
    Aplica formato profesional a las hojas de Excel
    columnas_ancho: dict con índices de columna y ancho personalizado
    formato_moneda: lista de índices de columnas a formatear como moneda
    """
    workbook = writer.book
    if sheet_name not in workbook.sheetnames:
        return
    
    ws = workbook[sheet_name]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    moneda_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Formato a encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Ancho de columnas (prioridad a anchos específicos)
    if columnas_ancho:
        for col_idx, ancho in columnas_ancho.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = ancho
    else:
        # Ancho automático basado en contenido
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Formato moneda y otros ajustes
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            if formato_moneda and idx in formato_moneda:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00'
                    cell.fill = moneda_fill
                    cell.alignment = center_alignment
            elif isinstance(cell.value, (int, float)):
                cell.alignment = center_alignment
            elif cell.value and isinstance(cell.value, str) and len(cell.value) > 50:
                cell.alignment = left_alignment

@app.route('/exportar/excel')
@login_required
def exportar_excel():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('tienda'))
    
    from models import ReporteETL
    import json
    
    # ============ 1. DATOS DEL ETL ============
    reporte_ventas = ReporteETL.query.filter_by(tipo_reporte='ventas').order_by(ReporteETL.fecha_reporte.desc()).first()
    reporte_productos = ReporteETL.query.filter_by(tipo_reporte='productos').order_by(ReporteETL.fecha_reporte.desc()).first()
    reporte_usuarios = ReporteETL.query.filter_by(tipo_reporte='usuarios').order_by(ReporteETL.fecha_reporte.desc()).first()
    
    ventas = json.loads(reporte_ventas.datos) if reporte_ventas else {}
    productos = json.loads(reporte_productos.datos) if reporte_productos else {}
    usuarios = json.loads(reporte_usuarios.datos) if reporte_usuarios else {}
    
    # ============ 2. DATOS EN VIVO DE LA BD ============
    with app.app_context():
        usuarios_completos = Usuario.query.all()
        df_usuarios_completo = pd.DataFrame([{
            'ID': u.id,
            'Nombre': u.nombre,
            'Usuario': u.usuario,
            'Email': u.email,
            'Celular': u.celular,
            'Rol': 'Administrador' if u.es_admin else 'Usuario',
            'Fecha Registro': u.fecha_registro.strftime('%Y-%m-%d') if u.fecha_registro else '',
            'Total Gastado': sum(o.total for o in u.ordenes) if u.ordenes else 0,
            'Número Órdenes': len(u.ordenes) if u.ordenes else 0
        } for u in usuarios_completos])
        
        productos_completos = Producto.query.all()
        df_productos_completo = pd.DataFrame([{
            'ID': p.id,
            'Nombre': p.nombre,
            'Descripción': (p.descripcion[:100] + '...') if p.descripcion and len(p.descripcion) > 100 else (p.descripcion or ''),
            'Precio': p.precio,
            'Stock': p.stock,
            'Categoría': p.categoria or 'Sin categoría',
            'Valor Inventario': p.precio * p.stock,
            'Fecha Creación': p.fecha_creacion.strftime('%Y-%m-%d') if p.fecha_creacion else ''
        } for p in productos_completos])
        
        ordenes_completas = Orden.query.all()
        df_ordenes = pd.DataFrame([{
            'ID Orden': o.id,
            'Cliente': o.usuario.nombre,
            'Usuario': o.usuario.usuario,
            'Fecha': o.fecha.strftime('%Y-%m-%d'),
            'Total': o.total,
            'Estado': o.estado
        } for o in ordenes_completas])
        
        items = OrdenItem.query.all()
        df_items = pd.DataFrame([{
            'Orden ID': i.orden_id,
            'Producto': i.producto.nombre,
            'Cantidad': i.cantidad,
            'Precio Unitario': i.precio_unitario,
            'Subtotal': i.cantidad * i.precio_unitario
        } for i in items])
        
        ventas_por_usuario = []
        for u in usuarios_completos:
            total = sum(o.total for o in u.ordenes) if u.ordenes else 0
            if total > 0:
                ventas_por_usuario.append({
                    'Usuario': u.usuario,
                    'Nombre': u.nombre,
                    'Total Gastado': total,
                    'Número Órdenes': len(u.ordenes) if u.ordenes else 0
                })
        df_ventas_usuario = pd.DataFrame(ventas_por_usuario).sort_values('Total Gastado', ascending=False)
    
    # ============ 3. CREAR EXCEL CON FORMATO ============
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Resumen Ejecutivo
        resumen = pd.DataFrame([
            ['📊 MÉTRICAS GENERALES', ''],
            ['Ventas Totales', f"${ventas.get('total_ventas', 0):,.2f}"],
            ['Total Órdenes', ventas.get('total_ordenes', 0)],
            ['Total Productos', productos.get('total_productos', 0)],
            ['Total Usuarios', usuarios.get('total_usuarios', 0)],
            ['Promedio por Venta', f"${ventas.get('promedio_venta', 0):,.2f}"],
            ['Stock Total', productos.get('stock_total', 0)],
            ['Productos Stock Bajo', len(productos.get('stock_bajo', []))],
            ['Administradores', usuarios.get('administradores', 0)],
            ['Usuarios Normales', usuarios.get('usuarios_normales', 0)],
        ])
        resumen.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False, header=False)
        
        # Ventas por Mes
        if ventas.get('ventas_por_mes'):
            df_meses = pd.DataFrame(ventas['ventas_por_mes'].items(), columns=['Mes', 'Total Ventas'])
            df_meses.to_excel(writer, sheet_name='Ventas por Mes', index=False)
        
        # Stock Bajo
        if productos.get('stock_bajo'):
            df_stock_bajo = pd.DataFrame(productos['stock_bajo'], columns=['Producto', 'Stock'])
            df_stock_bajo.to_excel(writer, sheet_name='Stock Bajo', index=False)
        
        # Top Productos ETL
        if productos.get('top_mas_vendidos'):
            df_top_etl = pd.DataFrame(productos['top_mas_vendidos'][:10], columns=['Producto', 'Unidades Vendidas'])
            df_top_etl.to_excel(writer, sheet_name='Top Productos', index=False)
        
        # Usuarios
        df_usuarios_completo.to_excel(writer, sheet_name='Usuarios', index=False)
        
        # Productos
        df_productos_completo.to_excel(writer, sheet_name='Productos', index=False)
        
        # Órdenes
        df_ordenes.to_excel(writer, sheet_name='Órdenes', index=False)
        
        # Detalle Compras
        df_items.to_excel(writer, sheet_name='Detalle Compras', index=False)
        
        # Ventas por Usuario
        if not df_ventas_usuario.empty:
            df_ventas_usuario.to_excel(writer, sheet_name='Ventas por Usuario', index=False)
        
        # Top Compradores
        if ventas.get('top_compradores'):
            df_top_compradores = pd.DataFrame(ventas['top_compradores'][:10], columns=['Usuario', 'Total Gastado'])
            df_top_compradores.to_excel(writer, sheet_name='Top Compradores', index=False)
        
        # ===== APLICAR FORMATO A CADA HOJA =====
        # Hoja: Usuarios (ancho específico)
        aplicar_formato_excel(writer, 'Usuarios', 
        columnas_ancho={1: 5, 2: 25, 3: 15, 4: 30, 5: 15, 6: 15, 7: 12, 8: 15, 9: 15},
        formato_moneda=[8])
        
        # Hoja: Productos (anchos específicos - IMPORTANTE)
        aplicar_formato_excel(writer, 'Productos',
        columnas_ancho={1: 5, 2: 30, 3: 40, 4: 12, 5: 8, 6: 15, 7: 15, 8: 12},
        formato_moneda=[4, 7])
        
        # Hoja: Órdenes
        aplicar_formato_excel(writer, 'Órdenes',
        columnas_ancho={1: 8, 2: 25, 3: 15, 4: 12, 5: 12, 6: 12},
        formato_moneda=[5])
        
        # Hoja: Detalle Compras
        if 'Detalle Compras' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Detalle Compras',
          columnas_ancho={1: 10, 2: 30, 3: 10, 4: 15, 5: 15},
          formato_moneda=[4, 5])
        
        # Hoja: Ventas por Usuario
        if 'Ventas por Usuario' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Ventas por Usuario',
          columnas_ancho={1: 20, 2: 25, 3: 15, 4: 12},
          formato_moneda=[3])
        
        # Hoja: Top Compradores
        if 'Top Compradores' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Top Compradores',
          columnas_ancho={1: 20, 2: 15},
          formato_moneda=[2])
        
        # Hoja: Ventas por Mes
        if 'Ventas por Mes' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Ventas por Mes',
          columnas_ancho={1: 12, 2: 15},
          formato_moneda=[2])

        # Hoja: Stock Bajo
        if 'Stock Bajo' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Stock Bajo',
          columnas_ancho={1: 30, 2: 10})

        # Hoja: Top Productos
        if 'Top Productos' in writer.book.sheetnames:
         aplicar_formato_excel(writer, 'Top Productos',
          columnas_ancho={1: 30, 2: 15})
        
        # Hojas sin formato moneda pero con ancho automático
        for sheet in ['Resumen Ejecutivo', 'Stock Bajo', 'Top Productos']:
            if sheet in writer.book.sheetnames:
                aplicar_formato_excel(writer, sheet)
    
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'reporte_etl_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)